#!/usr/bin/env python3
"""
Materials Export CLI - Export materials data to Excel
Usage: python materials_export.py [options]
"""

import os
import sys
import json
import argparse
from datetime import datetime
from pathlib import Path

# Add parent directory to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(__file__))))

try:
    from mp_api.client import MPRester
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(json.dumps({
        "error": f"Required package not installed: {e}",
        "status": "error",
        "hint": "Run: pip install mp-api pandas openpyxl"
    }))
    sys.exit(1)


def serialize_object(obj):
    """Convert objects to JSON-serializable format"""
    if obj is None:
        return None
    if isinstance(obj, (str, int, float, bool)):
        return obj
    if isinstance(obj, (list, tuple)):
        return [serialize_object(item) for item in obj]
    if isinstance(obj, dict):
        return {str(k): serialize_object(v) for k, v in obj.items()}
    if hasattr(obj, 'tolist'):
        return obj.tolist()
    if hasattr(obj, 'as_dict'):
        return serialize_object(obj.as_dict())
    if hasattr(obj, '__dict__'):
        return serialize_object(obj.__dict__)
    return str(obj)


def process_material_doc(doc):
    """Process material document into flat dictionary"""
    doc_dict = serialize_object(doc)

    result = {
        'Material_ID': doc_dict.get('material_id', ''),
        'Formula': doc_dict.get('formula_pretty', ''),
        'Band_Gap_eV': doc_dict.get('band_gap', ''),
        'Energy_Above_Hull_eV_Atom': doc_dict.get('energy_above_hull', ''),
        'Is_Stable': doc_dict.get('is_stable', ''),
        'Is_Metal': doc_dict.get('is_metal', ''),
        'Formation_Energy_eV_Atom': doc_dict.get('formation_energy_per_atom', ''),
        'Density_g_cm3': doc_dict.get('density', ''),
        'Volume_A3': doc_dict.get('volume', ''),
        'N_Sites': doc_dict.get('nsites', ''),
    }

    # Handle elements list
    elements = doc_dict.get('elements', [])
    if isinstance(elements, list):
        result['Elements'] = ', '.join(str(e) for e in elements)
    else:
        result['Elements'] = str(elements) if elements else ''

    # Symmetry
    symmetry = doc_dict.get('symmetry', {})
    if isinstance(symmetry, dict):
        result['Space_Group_Symbol'] = symmetry.get('symbol', '')
        result['Space_Group_Number'] = symmetry.get('number', '')

        # Handle crystal_system enum
        cs = symmetry.get('crystal_system')
        if isinstance(cs, dict) and '_value_' in cs:
            result['Crystal_System'] = cs['_value_']
        elif isinstance(cs, str):
            result['Crystal_System'] = cs
        else:
            result['Crystal_System'] = ''

    return result


def create_comparison_excel(materials_data, output_path):
    """Create horizontal comparison Excel format"""
    if not materials_data:
        return

    # Define all properties to include
    properties = [
        ('Material_ID', 'Material_ID'),
        ('Formula', 'Formula'),
        ('Band_Gap_eV', 'Band_Gap_eV'),
        ('Energy_Above_Hull_eV_Atom', 'Energy_Above_Hull_eV_Atom'),
        ('Is_Stable', 'Is_Stable'),
        ('Is_Metal', 'Is_Metal'),
        ('Formation_Energy_eV_Atom', 'Formation_Energy_eV_Atom'),
        ('Density_g_cm3', 'Density_g_cm3'),
        ('Volume_A3', 'Volume_A3'),
        ('N_Sites', 'N_Sites'),
        ('Elements', 'Elements'),
        ('Space_Group_Symbol', 'Space_Group_Symbol'),
        ('Space_Group_Number', 'Space_Group_Number'),
        ('Crystal_System', 'Crystal_System'),
    ]

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Materials Comparison"

    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    property_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    property_font = Font(bold=True, size=11)
    property_alignment = Alignment(horizontal="left", vertical="center")

    data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Write header row (material names)
    ws.cell(1, 1, "属性").fill = header_fill
    ws.cell(1, 1).font = header_font
    ws.cell(1, 1).alignment = header_alignment
    ws.cell(1, 1).border = border

    for col_idx, mat in enumerate(materials_data, start=2):
        formula = mat.get('Formula', 'N/A')
        mat_id = mat.get('Material_ID', 'N/A')
        header_text = f"{formula} ({mat_id})"
        cell = ws.cell(1, col_idx, header_text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Write data rows
    for row_idx, (prop_display, prop_key) in enumerate(properties, start=2):
        # Property name in first column
        cell = ws.cell(row_idx, 1, prop_display)
        cell.fill = property_fill
        cell.font = property_font
        cell.alignment = property_alignment
        cell.border = border

        # Material values in subsequent columns
        for col_idx, mat in enumerate(materials_data, start=2):
            value = mat.get(prop_key, None)

            # Format value - convert to string first to avoid MPID comparison issues
            if value is None:
                display_value = 'N/A'
            elif isinstance(value, bool):
                display_value = str(value)
            elif isinstance(value, (int, float)):
                if isinstance(value, float):
                    display_value = f"{value:.4g}"
                else:
                    display_value = str(value)
            else:
                # Convert to string first
                str_value = str(value)
                display_value = str_value if str_value else 'N/A'

            cell = ws.cell(row_idx, col_idx, display_value)
            cell.alignment = data_alignment
            cell.border = border

    # Set column widths
    ws.column_dimensions['A'].width = 30  # Property column
    for col_idx in range(2, len(materials_data) + 2):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 25  # Material columns

    # Set row heights
    ws.row_dimensions[1].height = 30  # Header row
    for row_idx in range(2, len(properties) + 2):
        ws.row_dimensions[row_idx].height = 25

    # Freeze panes
    ws.freeze_panes = 'B2'

    # Save workbook
    wb.save(output_path)


def style_excel_workbook(ws, df):
    """Apply professional styling to Excel worksheet"""
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_font = Font(bold=True, size=11, color="000000")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin', color='B4B4B4'),
        right=Side(style='thin', color='B4B4B4'),
        top=Side(style='thin', color='B4B4B4'),
        bottom=Side(style='thin', color='B4B4B4')
    )

    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Style data cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = cell_alignment
            cell.border = thin_border

    # Column widths
    for col_idx, col_name in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_idx)
        max_length = len(str(col_name))
        for cell in ws[col_letter][1:ws.max_row]:
            if cell.value:
                cell_len = len(str(cell.value).split('\n')[0])
                max_length = max(max_length, min(cell_len, 50))
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 50)

    # Enable filter and freeze panes
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'B2'


def export_materials(args):
    """Export materials to Excel"""
    api_key = os.environ.get("MP_API_KEY")
    if not api_key:
        return {
            "error": "MP_API_KEY environment variable not set",
            "status": "error"
        }

    try:
        with MPRester(api_key) as mpr:
            search_params = {}

            # Build search parameters
            if args.material_id:
                search_params["material_ids"] = [args.material_id]
            elif args.material_ids:
                search_params["material_ids"] = [mid.strip() for mid in args.material_ids.split(",")]

            if args.formula:
                search_params["formula"] = args.formula
            if args.elements:
                search_params["elements"] = args.elements.split(",")
            if args.chemsys:
                search_params["chemsys"] = args.chemsys

            if args.band_gap_min is not None or args.band_gap_max is not None:
                bg_min = args.band_gap_min if args.band_gap_min is not None else 0
                bg_max = args.band_gap_max if args.band_gap_max is not None else 100
                search_params["band_gap"] = (bg_min, bg_max)

            if args.stable:
                search_params["is_stable"] = True
            if args.metal:
                search_params["is_metal"] = True
            if args.magnetic:
                search_params["is_magnetic"] = True

            fields = [
                "material_id", "formula_pretty", "band_gap", "energy_above_hull",
                "is_stable", "is_metal", "formation_energy_per_atom", "density",
                "volume", "nsites", "elements", "symmetry"
            ]
            search_params["fields"] = fields

            # Only add chunk parameters if not searching by specific material IDs
            if not (args.material_id or args.material_ids):
                search_params["num_chunks"] = 1
                search_params["chunk_size"] = args.limit

            docs = mpr.materials.summary.search(**search_params)

            if not docs:
                return {
                    "error": "No materials found matching criteria",
                    "status": "error"
                }

            # Process materials
            materials_data = [process_material_doc(doc) for doc in docs]

            # Generate output path
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_dir = Path(args.output_dir)
            output_dir.mkdir(parents=True, exist_ok=True)

            if args.output:
                filename = args.output if args.output.endswith('.xlsx') else f"{args.output}.xlsx"
            elif args.material_ids:
                # Use first material ID for filename when comparing multiple
                first_id = args.material_ids.split(",")[0].strip()
                filename = f"{first_id}_comparison_{timestamp}.xlsx"
            elif args.material_id:
                filename = f"{args.material_id}_{timestamp}.xlsx"
            elif args.formula:
                filename = f"{args.formula}_{timestamp}.xlsx"
            else:
                filename = f"materials_export_{timestamp}.xlsx"

            output_path = output_dir / filename

            # Check if this is a comparison (multiple materials with --material-ids)
            use_comparison_format = args.material_ids and len(materials_data) >= 2

            if use_comparison_format:
                # Use horizontal comparison format
                create_comparison_excel(materials_data, output_path)
            else:
                # Use traditional vertical format
                # Create Excel workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Materials Data"

                df = pd.DataFrame(materials_data)

                # Write data
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        cell_value = str(value) if value is not None and value != '' else ""
                        ws.cell(row=r_idx, column=c_idx, value=cell_value)

                # Apply styling
                style_excel_workbook(ws, df)

                # Set row height
                for row_idx in range(2, ws.max_row + 1):
                    ws.row_dimensions[row_idx].height = 40

                wb.save(output_path)

            return {
                "status": "success",
                "file_path": str(output_path),
                "num_materials": len(materials_data),
                "timestamp": datetime.now().isoformat()
            }

    except Exception as e:
        import traceback
        return {
            "error": str(e),
            "traceback": traceback.format_exc(),
            "status": "error",
            "timestamp": datetime.now().isoformat()
        }


def main():
    parser = argparse.ArgumentParser(
        description="Export Materials Project data to Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Export specific material
  python materials_export.py --material-id mp-149 --output silicon.xlsx

  # Export and compare multiple materials
  python materials_export.py --material-ids mp-390,mp-672 --output TiO2_CdS_comparison.xlsx

  # Export semiconductors
  python materials_export.py --band-gap-min 1.0 --band-gap-max 3.0 --stable --limit 20

  # Export magnetic materials
  python materials_export.py --elements Fe,O --magnetic --output magnetic_materials.xlsx
        """
    )

    parser.add_argument("--material-id", help="Material ID (e.g., mp-149)")
    parser.add_argument("--material-ids", help="Comma-separated Material IDs (e.g., mp-149,mp-390,mp-672)")
    parser.add_argument("--formula", help="Chemical formula (e.g., Si, Fe2O3)")
    parser.add_argument("--elements", help="Comma-separated elements (e.g., Li,Fe,O)")
    parser.add_argument("--chemsys", help="Chemical system (e.g., Li-Fe-O)")
    parser.add_argument("--band-gap-min", type=float, help="Minimum band gap (eV)")
    parser.add_argument("--band-gap-max", type=float, help="Maximum band gap (eV)")
    parser.add_argument("--stable", action="store_true", help="Only stable materials")
    parser.add_argument("--metal", action="store_true", help="Only metallic materials")
    parser.add_argument("--magnetic", action="store_true", help="Only magnetic materials")
    parser.add_argument("--limit", type=int, default=10, help="Maximum results (default: 10)")
    parser.add_argument("--output", "-o", help="Output filename (without path)")
    parser.add_argument("--output-dir", default="./output", help="Output directory (default: ./output)")

    args = parser.parse_args()

    if len(sys.argv) == 1:
        parser.print_help()
        sys.exit(0)

    result = export_materials(args)

    if result.get("status") == "error":
        print(f"❌ Error: {result.get('error')}")
        sys.exit(1)

    print(f"\n✅ Export successful!")
    print(f"📁 File: {result['file_path']}")
    print(f"📊 Materials: {result['num_materials']}")


if __name__ == "__main__":
    main()
